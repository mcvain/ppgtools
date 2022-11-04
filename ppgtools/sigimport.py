from openpyxl import Workbook, load_workbook
import csv
import os
import numpy as np
from ppgtools.biosignal import BioSignal
import xlrd
import copy

class EventMarker:
    def __init__ (self, t, label):
        self.t = t
        self.label = label
    def __repr__(self):
        return str(self.t) + ": " + self.label
    def trim(markers, start_time, end_time):
        if start_time < 0:
            raise ValueError("Start time must be non-negative")
        out = []
        for e in markers:
            if start_time <= e.t <= end_time:
                out.append(EventMarker(e.t - start_time, e.label))
        
        return out      
    def remove_connection_events(markers):
        output = []
        
        non_con = 0
        for e in markers:
            if (not e.label.startswith("Device disconnected at")) and not (e.label.startswith("Device reconnected at")):
                output.append(e)
                print(str(non_con) + ": " + str(e.label))
                non_con += 1
                
        return output

    
#Write percentage of file read to console        
last_percent_written = -10
def checkPercent(total_bytes_read, file_size):
    global last_percent_written
    
    percent_complete = total_bytes_read / file_size * 100
    
    if(int(percent_complete) % 10 == 0 and int(percent_complete) != last_percent_written):
        print(str(int(percent_complete)) + "% complete")
        last_percent_written = int(percent_complete)      


def importTattooData(directory, filename):
    #Clean up path. Only use forward slashes
    directory = directory.replace("\\", "/")
    directory += "/" + filename
    
    devices = [name for name in os.listdir(directory) if os.path.isdir(os.path.join(directory, name))]
    print(devices)
    
    sessionData = {}
    
    for d in devices:
        data = importTAT(directory + "/" + d + "/" + filename + "_" + d)
        markers = importEventMarkers(directory + "/" + d + "/" + filename + "_" + d)
        deviceData = {"Data" : data, "Markers" : markers}
        sessionData[d] = deviceData
        
    print("All tattoo files loaded. Data available from following devices: ")    
    for d in sessionData.keys():
        print(d)
        
    return sessionData
        

def importEventMarkers(loc):
    '''
    Function to load event markers from Pulse App

    Parameters
    ----------
    loc : string
        Directory.

    Returns
    -------
    array_like
        List of EventMarkers.

    '''
    loc += ".csv"
    
    try:
        with open(loc, newline='') as f:
            reader = csv.reader(f)
            data = np.array(list(reader))
        
        out = []
        
        for i in range (0, len(data)):
            out.append(EventMarker(float(data[i][0]), data[i][1]))
        return out               

        print("Sucessfully loaded event markers")
    except FileNotFoundError:
        print("No event markers found")
        return []


def importTAT(loc):
    loc += ".tat"
    
    try:    
        file = open(loc, "rb")        
    except IOError:
        print("Could not find file \"" + loc + "\"")
        
    file_size = os.path.getsize(loc)
    print("File \'" + str(loc) + "\' size: " + str(file_size / 1000) + " kB")

    total_bytes_read = 0        
    
    i = 1
    #Parse the header for signal information
    byte = file.read(4)
    header_bytes = int.from_bytes(byte, "big")
    
    #Check version number of file
    byte = file.read(1)
    print("TAT file version: " + str(int.from_bytes(byte, "big")))
    
    signals = []
    print("\nReading file header (size: " + str(header_bytes) + " bytes) for signal information")
    while i < header_bytes:
        #Index, name length, name, bytes per point, fs, bit resolution, signed/unsigned
        print()
        byte = file.read(1)
        index = int.from_bytes(byte, "big")
        print("Index: " + str(index))
        
        byte = file.read(1)
        name_len = int.from_bytes(byte, "big")
        
        byte = file.read(name_len)
        name = byte.decode("utf-8")
        print("Name: " +  name)
        
        byte = file.read(1)
        bpp = int.from_bytes(byte, "big")
        print("Bytes per point: " +  str(bpp))
        
        byte = file.read(4)
        fs = int.from_bytes(byte, "big")
        print("Sample rate: " +  str(fs))
        
        byte = file.read(1)
        bit_res = int.from_bytes(byte, "big")
        print("Bit resolution: " +  str(bit_res))
        
        byte = file.read(1)
        signed = True#bool(int.from_bytes(byte, "big"))
        print("Signed: " +  str(signed))
        
        byte = file.read(1)
        little_endian = bool(int.from_bytes(byte, "big"))
        print("Little Endian: " +  str(little_endian))
        
        i+= (10 + name_len)
        
        signals.append(BioSignal(index, name, bpp, fs, bit_res, signed, little_endian))
    
    total_bytes_read += header_bytes
    checkPercent(total_bytes_read, file_size)
    
    #Get the order of which the signals are in
    print("\nDetermining the package structure...")
    byte = file.read(2)
    order_bytes = int.from_bytes(byte, "big")
    signal_order = []
    i = 0
    while i < order_bytes:
        byte = file.read(1)
        next_sig = int.from_bytes(byte, "big")
        signal_order.append(next_sig)
        print(str(next_sig))
        i += 1
    
    total_bytes_read += order_bytes
    checkPercent(total_bytes_read, file_size)
    
    #Parse the raw data   
    print("\nParsing raw data...")  
       
    
    data_buffer = [[] for i in range(len(signals))]
    while True:
        #Go through each signal
        for j in signal_order:
            bytes_to_read = signals[j].bytes_per_point
            byte = file.read(bytes_to_read)
            if (not byte):
                file.close()
                
                #Convert python list to numpy array
                for k in range (0, len(signals)):
                    signals[k].data = np.asarray(data_buffer[k])
               
                print("Sucessfully loaded .tat file.\n")
                return signals
            
            
            data_buffer[j].append(int.from_bytes(byte, signals[j].getEndian(), signed = signals[j].signed))
            #signals[j].data.append(int.from_bytes(byte, signals[j].getEndian(), signed = signals[j].signed))
            
    
            total_bytes_read += bytes_to_read
            checkPercent(total_bytes_read, file_size)  
               
def importBIN(loc):
    '''
    DEPRECATION WARNING. Future tattoo files will use .tat format.
    Converts a Pulse App Bin file into a list of BioSignals

    Parameters
    ----------
    loc : string
        Directory.

    Returns
    -------
    signals : array_like
        List of BioSignals.

    '''
    loc += ".bin"
    
    try:    
        file = open(loc, "rb")        
    except IOError:
        print("Could not find file \"" + loc + "\"")
        
    file_size = os.path.getsize(loc)
    print("File \'" + str(loc) + "\' size: " + str(file_size / 1000) + " kB")
        
    
        
    total_bytes_read = 0        
    
    i = 0
    #Parse the header for signal information
    byte = file.read(4)
    header_bytes = int.from_bytes(byte, "big")

    signals = []
    print("\nReading file header (size: " + str(header_bytes) + " bytes) for signal information")
    while i < header_bytes:
        #Index, name length, name, bytes per point, fs, bit resolution, signed/unsigned
        print()
        byte = file.read(1)
        index = int.from_bytes(byte, "big")
        print("Index: " + str(index))
        
        byte = file.read(1)
        name_len = int.from_bytes(byte, "big")
        
        byte = file.read(name_len)
        name = byte.decode("utf-8")
        print("Name: " +  name)
        
        byte = file.read(1)
        bpp = int.from_bytes(byte, "big")
        print("Bytes per point: " +  str(bpp))
        
        byte = file.read(4)
        fs = int.from_bytes(byte, "big")
        print("Sample rate: " +  str(fs))
        
        byte = file.read(1)
        bit_res = int.from_bytes(byte, "big")
        print("Bit resolution: " +  str(bit_res))
        
        byte = file.read(1)
        signed = True#bool(int.from_bytes(byte, "big"))
        print("Signed: " +  str(signed))
        
        byte = file.read(1)
        little_endian = bool(int.from_bytes(byte, "big"))
        print("Little Endian: " +  str(little_endian))
        
        i+= (10 + name_len)
        
        signals.append(BioSignal(index, name, bpp, fs, bit_res, signed, little_endian))
    
    total_bytes_read += header_bytes
    checkPercent(total_bytes_read, file_size)
    
    #Get the order of which the signals are in
    print("\nDetermining the package structure...")
    byte = file.read(2)
    order_bytes = int.from_bytes(byte, "big")
    signal_order = []
    i = 0
    while i < order_bytes:
        byte = file.read(1)
        next_sig = int.from_bytes(byte, "big")
        signal_order.append(next_sig)
        print(str(next_sig))
        i += 1
    
    total_bytes_read += order_bytes
    checkPercent(total_bytes_read, file_size)
    
    #Parse the raw data   
    print("\nParsing raw data...")  
       
    
    data_buffer = [[] for i in range(len(signals))]
    while True:
        #Go through each signal
        for j in signal_order:
            bytes_to_read = signals[j].bytes_per_point
            byte = file.read(bytes_to_read)
            if (not byte):
                file.close()
                
                #Convert python list to numpy array
                for k in range (0, len(signals)):
                    signals[k].data = np.asarray(data_buffer[k])
                
                return signals
            
            
            data_buffer[j].append(int.from_bytes(byte, signals[j].getEndian(), signed = signals[j].signed))
            #signals[j].data.append(int.from_bytes(byte, signals[j].getEndian(), signed = signals[j].signed))
            
    
            total_bytes_read += bytes_to_read
            checkPercent(total_bytes_read, file_size)
            
def importCSV(loc, num_channels, starting_row = 0, starting_col = 0):
    '''
    Converts a CSV file into signals. 

    Parameters
    ----------
    loc : string
        File location.
    num_channels : int
        How many columns of data to import.

    Returns
    -------
    out : array
        DESCRIPTION.

    '''
    
    loc += ".csv"
    
    with open(loc, newline='') as f:
        reader = csv.reader(f)
        data = np.array(list(reader))
    
    out = np.empty((num_channels - starting_col, data.shape[0] - starting_row-1))
    
    for i in range (starting_row, len(data)-1):
        for j in range (starting_col, num_channels):
            print(data[i][j])
            try:
                out[j - starting_col, i - starting_row] = (float(data[i][j]))
            except IndexError:
                print("Index Error, i = " + str(i) + " j = " + str(j))
        
            
    print("Sucessfully loaded .csv file.\n")
    
    return out

def import_and_convert_CSV(loc, signal_settings, file_name = ''):
    '''
    Converts CSV into a BioSignal.
    Option to resave the CSV data into a .bin

    Parameters
    ----------
    loc : string
        File location.
    signal_settings : array_like of BioSignal
        An array of BioSignals with the signal settings as required.
    file_name : string, optional
        When not empty, a new .bin file will be created from the CSV data. The default is "".        

    Returns
    -------
    out : array_like of BioSignal
        Returns an array of BioSignals.

    '''
    
    data = importCSV(loc, len(signal_settings))
    out = []
    #out = np.ndarray(len(signal_settings))
    
    #Attach the numpy array of data to the BioSignals
    for i in range(0, len(signal_settings)):
        signal_settings[i].data = data[i]
        out.append(signal_settings[i])
        
    out = np.asarray(out)    
        
    if(len(file_name) > 0):
        #create a new directory with the file name
        print(loc)
        loc = loc[0:loc.rfind("\\")+1]
        print(loc)

        #If directory doesn't already exist, make the directory.
        if(not os.path.isdir(loc + "/" + file_name)):
            #Make sure that all forward slashes are back slashes
            loc = loc.replace('/', '\\')
            print(loc)
            path = os.path.join(loc[0:-1], file_name)
            print(path)
            os.mkdir(path)
        
        #Create a new file in the directory
        f = open(loc + "/" + file_name + "/" + file_name + ".bin", "wb")
        xs = b""#bytearray(0)
        bytes_in_header = 0
        
        for i in range(0, len(out)):
            name_length = len(out[i].name)
            bytes_in_header += (10 + name_length)
            #Index, name length, name, bytes per point, fs (4), bit resolution, signed/unsigned, big/little
            xs += (out[i].index.to_bytes(1, 'big'))
            xs += (name_length.to_bytes(1, 'big'))
            xs += bytes(out[i].name, 'utf-8')
            xs += (out[i].bytes_per_point.to_bytes(1, 'big'))
            xs += (out[i].fs.to_bytes(4, 'big'))
            xs += (out[i].bit_resolution.to_bytes(1, 'big'))
            signed = 0
            xs += (signed.to_bytes(1, 'big'))
            #xs += (out[i].signed.to_bytes(1, 'big'))
            xs += (out[i].little_endian.to_bytes(1, 'big'))
            
        
        #Write the size of the header
        f.write(bytes_in_header.to_bytes(4, "big"))                
        
        #Write the header
        f.write(xs)
        
        #Write the data sequence. Assume signals are equal length and alternate.
        #First, write the number of signals to be parsed.
        bytes_in_sequence = len(out)
        f.write(bytes_in_sequence.to_bytes(2, 'big'))
        
        #Write the signal order
        for i in range(0, len(out)):
            f.write(i.to_bytes(1, 'big'))
            
        #Since we may need to convert floating point to fixed point, calculate the range of the data
        min_dat = []
        max_dat = []
        for j in range(0, len(out)):
            min_dat.append(min(out[j].data))
            max_dat.append(max(out[j].data))
            
        #Write the data.
        for i in range(0, len(out[0].data)):
            for j in range(0, len(out)):
                out_dat = out[j].data[i]
                #We need to convert from floating point to fixed point!
                range_dat = max_dat[j] - min_dat[j]
                
                #Normalize from 0 to 2^bitres
                out_dat = int(np.round((out_dat - min_dat[j]) / range_dat * (2**out[j].bit_resolution - 1)))
                
                f.write(out_dat.to_bytes(out[j].bytes_per_point, out[j].getEndian()))
        
        f.close()
        
    
    return out

def importSD(loc, endian = "little"):
    '''
    Function to load in data from old BIN files stored on SD card.

    Parameters
    ----------
    loc : String
        File name and location.
    endian : String, optional
        Identifier for what endian the data is in. The default is "little".

    Returns
    -------
    dist_data_complete : array_like
        Array of arrays of distal data.
    prox_data_complete : array_like
        Array of arrays of proximal data.
    lf_data_complete : array_like
        Array of arrays of the raw phototransistor signal.

    '''
    
    
    loc += ".bin"
    
    try:    
        file = open(loc, "rb")        
    except IOError:
        print("Could not find file \"" + loc + "\"")
        
    file_size = os.path.getsize(loc)
    print("File \'" + str(loc) + "\' size: " + str(file_size / 1000) + " kB")
        
    byte = file.read(1)
    n = 0
    
    sensor_data = 0
    dist_data_complete = []
    prox_data_complete = []
    lf_data_complete = []
    
    prox_data = []
    dist_data = []
    lf_data = []
    
    total_bytes_read = 0
    last_percent_written = -10
    
    while byte:
        #First byte   
        if n % 2 == 0:
            sensor_data = byte
        else:
            sensor_data = int.from_bytes(sensor_data + byte, endian)
            #print(sensor_data)            
            if sensor_data == 0xFFFF:
                #A new boot
                if(len(dist_data) != 0):
                    dist_data_complete.append(dist_data)
                    prox_data_complete.append(prox_data)
                    lf_data_complete.append(lf_data)
                    dist_data = []
                    prox_data = []
                    lf_data = []
                n = -1
                
            else:
                if sensor_data & 0xFC00 == 0x400:
                    dist_data.append(sensor_data & 0x3FF)
                elif sensor_data & 0xFC00 == 0x800:
                    lf_data.append((sensor_data & 0x3FF))
                else:
                    prox_data.append(sensor_data & 0x3FF)
                
        #sys.stdout.write("-")
        #sys.stdout.flush()
        n += 1
        total_bytes_read += 1
        
        percent_complete = total_bytes_read / file_size * 100
        
        if(int(percent_complete) % 10 == 0 and int(percent_complete) != last_percent_written):
            print(str(int(percent_complete)) + "% complete")
            last_percent_written = int(percent_complete)
        
        byte = file.read(1)
    
    #sys.stdout.write("]\n")
    
    file.close()
        
    prox_data_complete.append(prox_data)
    dist_data_complete.append(dist_data)
    lf_data_complete.append(lf_data)
    
    print("")
    print(str(len(dist_data_complete)) + " datasets found in " + loc + "\n")
    
    return dist_data_complete, prox_data_complete, lf_data_complete    

def importWellueO2Data(path, loc = 'all'):        
    spo2 = []
    hr = []
    motion = []
    
    if loc == 'all':
        onlyfiles = [f for f in os.listdir(path) if os.path.isfile(os.path.join(path, f))]
        print("O2 files found: ")
        for i in range(0, len(onlyfiles)):
            print(f"{i}: {onlyfiles[i]}")
        locs = onlyfiles
    else:
        locs = [loc+".csv"]
    
    for l in locs:
        with open(path + '\\' + l, newline='', errors = 'ignore') as f:
            reader = csv.reader(f)
            data = np.array(list(reader))
            
        for i in range (1, len(data)):        
            spo2.append(float(data[i][1]))
            hr.append(float(data[i][2]))
            motion.append(float(data[i][3]))
    
    return list((spo2, hr, motion))

def importBiopacEventMarkers(loc):
    book = xlrd.open_workbook(loc)
    sh = book.sheet_by_index(0)
    
    markers = []
    beats = []
    ibi = []
        
    for row in range (2, sh.nrows):        
        t = sh.cell(row, 1).value
        
        if t[-4:] != ' sec':
            print(t)
        else:
            t = float(t[0:-4])
        
        event = sh.cell(row, 2).value
        
        n = 0
        if event == "User Type 1":
            print(f"{len(markers)}: {sh.cell(row, 4).value}") 
            markers.append(EventMarker(t, sh.cell(row, 4).value))
        elif event == "Normal Beat":
            beats.append(t)
        elif event == "Premature Ventricular Contraction":
            beats.append(t)
            markers.append(EventMarker(t, "PVC"))
        else:
            print(event) 
            
    beats = np.asarray(beats)
    markers = np.asarray(markers)
    ibi = np.diff(beats)
    hr = 60 / ibi
                  
    #Remove the last beat since there is 1 less IBI than beats
    return beats[0:-1], ibi, hr, markers

def importBiopacRespiratoryData(loc):
    breaths = []

    book = xlrd.open_workbook(loc)
    sh = book.sheet_by_index(0)
        
    for row in range (2, sh.nrows):        
        t = sh.cell(row, 1).value
        
        if t[-4:] != ' sec':
            print(t)
        else:
            t = float(t[0:-4])
        
        event = sh.cell(row, 2).value
        
        if event == "Inspire Start":
            breaths.append(t)
        elif event == "Expire Start":
            pass
        elif event == "Recovery":
            pass
        else:
            pass
            #print(event) 
            
    breaths = np.asarray(breaths)
    ibi = np.diff(breaths)
    rr = 60 / ibi
                  
    #Remove the last beat since there is 1 less IBI than beats
    return breaths[0:-1], ibi, rr
        

def importFinapres(loc):
    '''
    Deprecated. Want to update so it returns BioSignal.

    Parameters
    ----------
    loc : string
        File location.

    Returns
    -------
    time : array_like
        Timestamps of the BP data.
    bp : array_like
        Continuous BP data.

    '''
    loc += ".csv"
    
    time = []
    bp = []
    with open(loc, newline='') as f:
        reader = csv.reader(f)
        data = list(reader)
    
    for i in range (2, len(data)):
        cur_bp = float(data[i][1])
        
        if(cur_bp > 30):
            time.append(float(data[i][0]))
            bp.append(float(data[i][1]))
            
    print("Sucessfully loaded BP data")
    
    return time, bp

def generateSine(length, fs, freq, amplitude = 1, delay = 0):
    '''
    Function to generate a BioSignal containing a sine wave.

    Parameters
    ----------
    length : int
        Length of the data in seconds.
    fs : int
        Simulated frequency of the sine wave.
    freq : int
        Frequency of the sine wave.
    amplitude : float, optional
        Amplitude of the sine wave. The default is 1.
    delay : float, optional
        Delay the signal by this in ms. The default is 0.

    Returns
    -------
    output : TYPE
        DESCRIPTION.

    '''
    output = BioSignal(0, "f = " + str(freq) + " Hz", 0, fs, 0, True, False)
    t = np.linspace(0, length, length*fs)
    output.data = amplitude * np.sin(2*np.pi*(t+delay))
    
    return output
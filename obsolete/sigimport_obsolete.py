from openpyxl import workbook, load_workbook
import csv
import os
import numpy as np
    
def importXLSX(loc):
    '''
    Obsolete. Do not use.

    Parameters
    ----------
    loc : TYPE
        DESCRIPTION.

    Returns
    -------
    list
        DESCRIPTION.
    list
        DESCRIPTION.
    list
        DESCRIPTION.

    '''
    if loc.endswith('.xlsx') == False:
        loc += ".xlsx"
    
    book = load_workbook(filename = loc)
    print(book.sheetnames)
    sheet = book.active
    
    channel0 = []
    channel1 = []
    channel2 = []
    
    print(sheet.max_row)
    print(sheet.max_column)
    for r in range(1, sheet.max_row):
        channel0.append(sheet.cell(r, 1).value)
        if sheet.max_column >=3:
            channel1.append(sheet.cell(r, 2).value)
        if sheet.max_column >=4:
            channel2.append(sheet.cell(r, 3).value)
        

    print("Sucessfully loaded file")
    
    return [channel0], [channel1], [channel2]
    
    
def importBIN(loc, endian = "little"):
    '''
    Obsolete. Do not use.

    Parameters
    ----------
    loc : TYPE
        DESCRIPTION.
    endian : TYPE, optional
        DESCRIPTION. The default is "little".

    Returns
    -------
    dist_data_complete : TYPE
        DESCRIPTION.
    prox_data_complete : TYPE
        DESCRIPTION.
    lf_data_complete : TYPE
        DESCRIPTION.

    '''
    loc += ".bin"
    
    try:    
        file = open(loc, "rb")        
    except IOError:
        print("Could not find file \"" + loc + "\"")
        
    file_size = os.path.getsize(loc)
    print("File \'" + str(loc) + "\' size: " + str(file_size / 1000) + " kB")
        
    byte = file.read(1)
    n = 0
    
    sensor_data = 0
    dist_data_complete = []
    prox_data_complete = []
    lf_data_complete = []
    
    prox_data = []
    dist_data = []
    lf_data = []
    
    total_bytes_read = 0
    last_percent_written = -10
    
    while byte:
        #First byte   
        if n % 2 == 0:
            sensor_data = byte
        else:
            sensor_data = int.from_bytes(sensor_data + byte, endian)
            #print(sensor_data)            
            if sensor_data == 0xFFFF:
                #A new boot
                if(len(dist_data) != 0):
                    dist_data_complete.append(dist_data)
                    prox_data_complete.append(prox_data)
                    lf_data_complete.append(lf_data)
                    dist_data = []
                    prox_data = []
                    lf_data = []
                n = -1
                
            else:
                if sensor_data & 0xFC00 == 0x400:
                    dist_data.append(sensor_data & 0x3FF)
                elif sensor_data & 0xFC00 == 0x800:
                    lf_data.append((sensor_data & 0x3FF))
                else:
                    prox_data.append(sensor_data & 0x3FF)
                
        #sys.stdout.write("-")
        #sys.stdout.flush()
        n += 1
        total_bytes_read += 1
        
        percent_complete = total_bytes_read / file_size * 100
        
        if(int(percent_complete) % 10 == 0 and int(percent_complete) != last_percent_written):
            print(str(int(percent_complete)) + "% complete")
            last_percent_written = int(percent_complete)
        
        byte = file.read(1)
    
    #sys.stdout.write("]\n")
    
    file.close()
        
    prox_data_complete.append(prox_data)
    dist_data_complete.append(dist_data)
    lf_data_complete.append(lf_data)
    
    print("")
    print(str(len(dist_data_complete)) + " datasets found in " + loc + "\n")
    
    return dist_data_complete, prox_data_complete, lf_data_complete


# =============================================================================
# Script to combine multiple phone recordings to a single BIN file, iterating based on a common part of the file name
# =============================================================================
def combine_BINs_numerated(loc, end_num, output_name, end_string = "", start_num = 1):
    '''
    Obsolete. Do not use.

    Parameters
    ----------
    loc : TYPE
        DESCRIPTION.
    end_num : TYPE
        DESCRIPTION.
    output_name : TYPE
        DESCRIPTION.
    end_string : TYPE, optional
        DESCRIPTION. The default is "".
    start_num : TYPE, optional
        DESCRIPTION. The default is 1.

    Returns
    -------
    None.

    '''
    f_out = open(loc + "\\" + output_name + ".bin", 'wb')
    succ_files = []
    
    for i in range (start_num, end_num + 1):
        cur_file = loc + "\\" + str(i) + end_string +  ".bin"
    
        try:    
            f_in = open(cur_file, "rb")    
            file_size = os.path.getsize(cur_file)
            print("File \'" + str(cur_file) + "\' size: " + str(file_size / 1000) + " kB")
            f_out.write(f_in.read())
            f_in.close
            succ_files.append(str(i) + end_string)
        except IOError:
            print("Could not find file \"" + cur_file + "\"")

    f_out.close
    print("Successful files: ")
    for i in range(0, len(succ_files)):
        print(succ_files[i])
        
# =============================================================================
# Script to combine multiple phone recordings to a single BIN file, from a list of file names
# =============================================================================
def combine_BINs_list(loc, file_list, output_name):
    '''
    Obsolete. Do not use.

    Parameters
    ----------
    loc : TYPE
        DESCRIPTION.
    file_list : TYPE
        DESCRIPTION.
    output_name : TYPE
        DESCRIPTION.

    Returns
    -------
    None.

    '''
    f_out = open(loc + "\\" + output_name + ".bin", 'wb')
    failed_files = []
    
    for i in range (0, len(file_list)):
        cur_file = loc + "\\" + file_list[i] +  ".bin"
    
        try:    
            f_in = open(cur_file, "rb")    
            file_size = os.path.getsize(cur_file)
            print("File \'" + str(cur_file) + "\' size: " + str(file_size / 1000) + " kB")
            f_out.write(f_in.read())
            f_in.close
        except IOError:
            print("Could not find file \"" + cur_file + "\"")
            failed_files.append(file_list[i])

    f_out.close
    print("Failed files: ")
    for i in range(0, len(failed_files)):
        print(failed_files[i])        

# =============================================================================
# Script to combine multiple Labview (.xlsx) files to a single BIN file, iterating based on a common part of the file name
# For now, this only converts the first two channels        
# =============================================================================
def combine_XLSXs_numerated(loc, end_num, output_name, end_string = "", start_num = 1):
    '''
    Obsolete. Do not use.

    Parameters
    ----------
    loc : TYPE
        DESCRIPTION.
    end_num : TYPE
        DESCRIPTION.
    output_name : TYPE
        DESCRIPTION.
    end_string : TYPE, optional
        DESCRIPTION. The default is "".
    start_num : TYPE, optional
        DESCRIPTION. The default is 1.

    Returns
    -------
    None.

    '''
    f_out = open(loc + "\\" + output_name + ".bin", 'wb')
    succ_files = []
    
    for i in range (start_num, end_num + 1):
        cur_file = loc + "\\" + str(i) + end_string +  ".xlsx"
        
        in0 = []
        in1 = []
        in2 = []
    
        try:    
            in0, in1, in2 = importXLSX(cur_file)
            in0 = in0[0]
            in1 = in1[0]
            
            if(len(in2) != 0):
                in2=in2[0]
            
            f_out.write(0xFFFF.to_bytes(2, byteorder = 'little'))
            for j in range (1, len(in0)):
                out0 = ((int(float(in0[j]) * 65536) >> 8) | 0x400).to_bytes(2, byteorder = 'little')
                out1 = (int(float(in1[j]) * 65536) >> 8).to_bytes(2, byteorder = 'little')
                
                #print(str((int(float(in0[j]) * 65536) >> 6) | 0x400) + '\t' + str((int(float(in1[j]) * 65536) >> 6)))
                
                
                f_out.write(out0)
                f_out.write(out1)
                
                if(len(in2) != 0):
                    out2 = ((int(float(in2[j]) * 65536) >> 8) | 0x800).to_bytes(2, byteorder = 'little')
                    f_out.write(out2)
            
            
            succ_files.append(str(i) + end_string)
        except IOError:
            print("Could not find file \"" + cur_file + "\"")

    f_out.close
    print("Successful files: ")
    for i in range(0, len(succ_files)):
        print(succ_files[i])        